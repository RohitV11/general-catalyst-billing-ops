"""
NCCI Procedure-to-Procedure (PTP) edit constraints.

Loads CMS Correct Coding Initiative edit tables from the quarterly
Practitioner PTP Excel files in data/ and exposes two data structures
plus one lookup function for billing-compliance checks.

Data structures
---------------
HARD_BLOCKS : set[tuple[str, str]]
    Code pairs that can NEVER be billed together (modifier indicator = 0).
MODIFIER_ALLOWED : dict[tuple[str, str], str]
    Code pairs that CAN be billed together when the appropriate modifier
    (e.g. "59" — distinct procedural service) is appended.
    The value is the modifier indicator from the edit table ("1").

Function
--------
check_pair(code1, code2) -> dict
    Returns whether two CPT codes can be billed together and what
    modifier is required, if any.
"""

import os
import glob
import openpyxl  # pyre-ignore[21]

# load PTP files

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data")
HEADER_ROWS = 6 

HARD_BLOCKS = set()
MODIFIER_ALLOWED = {}


def load_edits():
    xlsx_files = sorted(glob.glob(os.path.join(DATA_DIR, "*.xlsx")))
    if not xlsx_files:
        raise FileNotFoundError(f"No .xlsx files found in {DATA_DIR}")
        
    for fpath in xlsx_files:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # print column names to catch mismatches
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 3:
                header_row = [str(c) if c else "" for c in row]
                print(f"[constraints] {os.path.basename(fpath)} columns: {header_row}")
            if i <= HEADER_ROWS:
                continue

            col1, col2, _exist, _eff_date, del_date, modifier, _rationale = row

            if col1 is None:
                continue

            del_str = str(del_date).strip() if del_date is not None else "*"
            if del_str != "*":
                continue

            pair = (str(col1).strip(), str(col2).strip())
            mod = str(modifier).strip() if modifier is not None else ""

            if mod == "0":
                # indicates hard blocks are never allowed together 
                HARD_BLOCKS.add(pair)
            elif mod == "1":
                # indicates pairs that are allowed with a specific modifier
                MODIFIER_ALLOWED[pair] = mod

        wb.close()


load_edits()


def check_pair(code1, code2):
    code1, code2 = code1.strip(), code2.strip()
    pair = (code1, code2)
    pair_rev = (code2, code1)

    # check the hard blocks first
    if pair in HARD_BLOCKS or pair_rev in HARD_BLOCKS:
        return {
            "allowed": False,
            "modifier": None,
            "reason": (
                f"CCI edit: {code1} and {code2} can NEVER be billed together "
                "(modifier indicator 0)."
            ),
        }

    # then check modifier-allowed edits in case no hard blocks are violated
    if pair in MODIFIER_ALLOWED:
        return {
            "allowed": True,
            "modifier": MODIFIER_ALLOWED[pair],
            "reason": (
                f"CCI edit: {code1} and {code2} can be billed together with "
                "an appropriate modifier (e.g. modifier 59 — distinct procedural service)."
            ),
        }
    if pair_rev in MODIFIER_ALLOWED:
        return {
            "allowed": True,
            "modifier": MODIFIER_ALLOWED[pair_rev],
            "reason": (
                f"CCI edit: {code2} and {code1} can be billed together with "
                "an appropriate modifier (e.g. modifier 59 — distinct procedural service)."
            ),
        }

    # if no edit file exists, that means there are no restriction and these 2 codes can be billed together
    return {
        "allowed": True,
        "modifier": None,
        "reason": (
            f"No CCI edit exists for {code1} and {code2}. "
            "They can be billed together without restrictions."
        ),
    }


if __name__ == "__main__":
    print(f"\nHard-blocked pairs:       {len(HARD_BLOCKS):>10,}")
    print(f"Modifier-allowed pairs:   {len(MODIFIER_ALLOWED):>10,}")
    print(f"Total constrained pairs:  {len(HARD_BLOCKS) + len(MODIFIER_ALLOWED):>10,}")
    print()

    # example lookups
    examples = [("99213", "99214"), ("76700", "76705")]
    for c1, c2 in examples:
        result = check_pair(c1, c2)
        status = "ALLOWED" if result["allowed"] else "BLOCKED"
        print(f"  {c1} + {c2}: {status} — {result['reason']}")



